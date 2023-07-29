import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QTableWidgetItem, QFileDialog, QDialog, QMessageBox
from PyQt5.QtCore import pyqtSlot, QFile, QTextStream, pyqtSignal
from ELIZADE_ui import Ui_MainWindow
import openpyxl
from scraper2 import scrape
import sqlite3
import os
from Results import Results
import sidebar_main
import departmental
import senate
import student
# from randomNew import MyWindow 
# from PyQt5.uic import loadUi
from login_ui import Ui_Dialog

def validate_credentials(username, password):
    if username == "uname" and password == "Password":
        return "Students"
    elif username =="Username" and password == "Password":
        return "lecturer"
    else:
        return None


class Login(QDialog):
    login_successful = pyqtSignal(str)
    def __init__(self):
        super(Login, self).__init__()

        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        self.ui.loginbutton.clicked.connect(self.loginfunction)
        self.ui.password.setEchoMode(QtWidgets.QLineEdit.Password)
        

    def loginfunction(self):
        username = self.ui.username.text()
        password = self.ui.password.text()

        # Validate the login credentials and check in the database
        conn = sqlite3.connect("main.db")
        cursor = conn.cursor()

        # Check against Lecturers table
        cursor.execute("SELECT * FROM Lecturers WHERE Username = ? AND Password=?", (username, password))
        user_Lecturers = cursor.fetchall()
        # print(f"Here: {user_Lecturers}") 

        # Check against Students table
        cursor.execute("SELECT * FROM Students WHERE uname = ? AND Password = ?", (username, password))
        user_Students = cursor.fetchall()
        # print(f"Student: {user_Students}")

        if user_Lecturers or user_Students:
            QMessageBox.information(self, 'Login', 'Login successful!')
            #self.main_window.set_username(username) #notify the main window and rerturn sucess signal
            # self.accept()  # Close the login window and return success signal

            if user_Lecturers:
                login_type = 'lecturer'

            else:
                login_type = 'student'
            self.login_successful.emit(login_type)
            self.accept()
        else:
            QMessageBox.warning(self, 'Login', 'Invalid credentials.')
            print("Login failed. Incorrect username or password.")

        conn.close()





class LecturerWindow(QMainWindow):
    def __init__(self):
        super(LecturerWindow, self).__init__()
        
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.icon_only_widget.hide()
        self.ui.stackedWidget.setCurrentIndex(0)
        self.ui.home_btn_2.setChecked(True)
        self.uploadresult()

        self.headerList = ['S/N', "Name", "Matric No.", "Grade", 'TCP', 'TNU', 'GPA', 'Remarks']
        self.ui.btn_prev.clicked.connect(lambda: senate.show_data(self, -10, 0, 'students.db'))
        self.ui.btn_next.clicked.connect(lambda: senate.show_data(self, 10, 0, 'students.db'))
        self.ui.btn_mat.clicked.connect(lambda: senate.show_data(self, 0, 1, 'students.db'))
        self.ui.btn_gpa.clicked.connect(lambda: senate.show_data(self, 0, 2, 'students.db'))

        self.ui.search.clicked.connect(self.browseFiles)
        self.ui.load.clicked.connect(self.load_data)
        self.ui.save.clicked.connect(self.save_data)

        self.offset = 0
    
        # Initialize login dialog and pass self as main_window argument
        # self.login_dialog = Login()
        # self.login_dialog.show()

    def set_username(n):
        if n is not None and len(n)> 0: #check might not be required if you;re okay with empty names
            self.name = n

# in function that creates login Window, create the login window with a reference
  # window creation might already take a "parent" param, in which case you're all set.
        # createLoginWindow(requiredParams, main=self) 

# in LoginWindow init/constructor:
        # self.mainWindow = main

# When name changes:
        self.mainWindow.setName(name)
    ## function for searching
    def on_search_btn_clicked(self): 
        self.ui.stackedWidget.setCurrentIndex(10)
        search_text = self.ui.search_input.text().strip()
        if search_text:
            self.ui.label_11.setText(search_text) 

    ## function for changing page to user page
    def on_user_btn_clicked(self):
        self.ui.stackedWidget.setCurrentIndex(11)
    def temporary(self, data):
        if data == "First Semester":
            return 1
        else:
            return 2
    def save_data(self):
        entry1 = self.ui.combo3.currentText()
        entry2 = self.temporary(self.ui.combo2.currentText())
        entry3 = self.ui.combo1.currentText()
        rese = Results(entry1, entry2, entry3)
        rese.extract_data_and_insert()
    
    def browseFiles(self):
        fname=QFileDialog.getOpenFileName(self, 'open file', '100l.html')
        self.ui.filename.setText(fname[0])
    
    def uploadresult(self):
        self.ui.tableWidget.setRowCount(23)
        self.ui.tableWidget.setColumnCount(14)
    
    def load_data(self):    
        
        path = self.ui.filename.text()
        res = scrape(path)
        workbook = openpyxl.load_workbook(res)
        sheet = workbook.active
        workbook.save(res)

        self.ui.tableWidget.setRowCount(sheet.max_row)
        self.ui.tableWidget.setColumnCount(sheet.max_column)
        list_values = list(sheet.values)

        self.ui.tableWidget.setHorizontalHeaderLabels(list_values[0])
        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.ui.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1 
        for row in range(self.ui.tableWidget.rowCount()):
            for col in range(self.ui.tableWidget.columnCount()):
                cell = self.ui.tableWidget.item(row, col)
                if cell is not None and cell.text() == 'None':
                    cell.setText('-')

    ## Change QPushButton Checkable status when stackedWidget index change
    def on_stackedWidget_currentChanged(self, index):
        btn_list = self.ui.icon_only_widget.findChildren(QPushButton) \
                    + self.ui.full_menu_widget.findChildren(QPushButton)
        
        for btn in btn_list:
            if index in [10, 11]:
                btn.setAutoExclusive(False)
                btn.setChecked(False)
            else:
                btn.setAutoExclusive(True)


    ## function for changing menu page
    def on_home_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(0)

    def on_home_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(0)

    def on_dashboard_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(1)

    def on_dashboard_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(1)

    def on_result_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(2)

    def on_departmental_btn_1_toggled(self):
        connection = sqlite3.connect("main.db")
        print(connection)
        cursor = connection.cursor()

        self.ui.stackedWidget.setCurrentIndex(2)
        self.departmental = departmental.display_departmental_summary(self, cursor)
        # self.departmental.show()    
        connection.close()

    def on_alumni_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(3)

    def on_student_btn_3_toggled(self):
        connection = sqlite3.connect("main.db")
        cursor = connection.cursor()

        self.ui.stackedWidget.setCurrentIndex(4)
        self.student = student.display_senate_summary(self, 'EU210303-2874', cursor)
        
        connection.close()

    def on_senate_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(5)
        self.senate = senate.show_data(self, 0, 0, 'students.db ')

    def on_registration_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(6)


    def on_registration_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(6)

    def on_upload_result_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(7)

    def on_upload_result_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(7)

    def on_curriculum_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(8)

    def on_curriculum_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(8)

    def on_student_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(9)

    def on_student_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(9)

    def on_search_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(11)

    def on_search_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(11)

    def on_user_btn_1_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(8)

    def on_user_btn_2_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(8)
        

    def set_username(self,username):
        self.username = username

class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()


        self.login_dialog = Login()
        self.login_dialog.login_successful.connect(self.on_login_successful)
        self.login_dialog.show()
        res = self.login_dialog.exec_()
        # self.type = self.login_dialog.type
        self.setCentralWidget(self.login_dialog)
    
    def on_login_successful(self, login_type):
        # This slot will be called when the login is successful
        print(f"Login successful. Type: {login_type}")

        # Do whatever you want with the login_type (lecturer or student).
        # For example, you can use it to determine which interface to show:
        if login_type == 'lecturer':
            # Show the lecturer interface
            self.lecturer_window = LecturerWindow()
            # self.setCentralWidget(self.lecturer_window)
            self.lecturer_window.show()
        elif login_type == 'student':
            # Show the student interface
            print("Here")
            self.student_window = sidebar_main.MainWindow()  # Assuming there's a StudentWindow class for the student interface
            self.student_window.show()
            # self.setCentralWidget(self.student_window)
        
    
   

if __name__ == "__main__":
    app = QApplication(sys.argv)
    

    # loading style file
    with open("style.qss", "r") as style_file:
        style_str = style_file.read()
    app.setStyleSheet(style_str)

    # main_window = LecturerWindow()
    # main_window.login_dialog.show()  # Show the login dialog
    # main_window.show()
    # login_dialog = Login()
    # login_dialog.show()

    # result = login_dialog.exec_()
    # if result == QDialog.Accepted:
        # print("Login successful. Showing the main window.")
        # main_window.show()
        
        # sys.exit(app.exec_())
    # print('Ënd 1')

    # if main_window.login_dialog.exec_() == 1:
    #     print("End 2")
    main_window = MainWindow()
    # main_window.show()
    sys.exit(app.exec_())